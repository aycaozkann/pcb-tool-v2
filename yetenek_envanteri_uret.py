#!/usr/bin/env python3
"""
yetenek_envanteri_uret.py
===========================
Repo'daki her modülü tarayıp otomatik güncellenen bir "Yetenek Envanteri"
(YETENEK_ENVANTERI.xlsx) üretir — "bu araç şunu yapabiliyor mu" sorusuna
elle repo taramak yerine tek bir Excel dosyasına bakarak cevap vermek için.

NEDEN BU DOSYA VAR:
--------------------
Bu proje büyüdükçe (70+ Python modülü) hangi modülün TAM işlevsel, hangisinin
İSKELET, hangisinin gerçekten test edildiği sorusu sohbet geçmişine/hafızaya
gömülü kalıyordu. Bu modül o analiz tablosunu KOD TARAFINDAN, repo'nun
gerçek dosyalarını okuyarak üretir.

DÜRÜSTLÜK SINIRI (bilerek, defalarca altı çizilen bir disiplin):
------------------------------------------------------------------
Bu envanter mimari kararları OTOMATİK VERMEZ — sadece "şu an ne var, ne
yok, ne yarım" sorusuna hızlı, güncel bir cevap sağlar:
  - `Durum` (Var/Kısmen var/İskelet) markerlı bir docstring bulunamazsa
    fonksiyon sayısı/gövde uzunluğuna göre KABA bir TAHMİNDİR — `durum_
    kaynagi="otomatik_tahmin"` ile İŞARETLENİR, kesin karar İDDİA EDİLMEZ.
  - `gercek_mi_mock_mu` sadece ŞÜPHELİ desenleri (mock/sahte/taklit
    kelimeleri, tekrarlanan sabit-değer-dönen if/elif zincirleri) işaretler
    — bir modülü otomatik olarak "mock" diye MAHKÛM ETMEZ, sadece
    "KONTROL EDİLMELİ" bayrağı koyar.
  - `arac_zinciri_uyumu` (Tam/Kısmen uyumlu, İlgisiz) json+döngü deseni
    aramasıyla YAKLAŞIK işaretlenir, kesin bir statik analiz DEĞİLDİR.
  - `test_durumu` TEK istisna — bu GERÇEKTEN `pytest --junit-xml` ile
    ölçülür (bkz. `tum_testleri_calistir_ve_sonuclari_topla()`), sabit/
    varsayılan bir "hepsi geçti" UYDURULMAZ; testleri atlamak isteyen
    çağıran `--testleri-atla` ile bunu AÇIKÇA seçmelidir.

Kritik bir karar öncesi ("bu özelliği kullanabilir miyiz") yine de ilgili
dosyanın ELLE kontrol edilmesi gerekebilir — bu, Excel'in Özet sayfasında
da bir hatırlatma notu olarak bulunur.

OTOMATİK GÜNCELLEME: `main.py envanter-guncelle` CLI bayrağı bu modülü
çağırır (bkz. `main.py::cmd_envanter_guncelle`). Bir git hook'una BİLEREK
BAĞLANMADI — bu projenin genel disiplini (bu oturumda defalarca uygulandı)
her adımın kod+test+suite+COMMIT'inin elle/bilinçli yapılmasıdır; otomatik
bir git-commit hook'u bu disiplinle çelişir. `KURULUM.md`'ye bu CLI
bayrağının oturum sonunda elle çağrılması gerektiği not düşüldü.
"""

from __future__ import annotations

import argparse
import ast
import re
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

def turkce_kucuk_harfe_cevir(metin: str) -> str:
    """Python'un `str.lower()`'ı Türkçe `İ` (noktalı büyük I) harfini `i` +
    BİRLEŞTİRİCİ NOKTA (U+0307) olarak çevirir (Unicode'un locale-bağımsız
    kuralı) — bu hem basit alt-dize eşleşmesini SESSİZCE kırar (bu dosyanın
    testinde bulundu: 'DEĞİL'.lower() 'değil' İLE EŞLEŞMEZ) HEM DE Windows
    konsolunda (cp125x) `UnicodeEncodeError` fırlatabilir. Türkçe metin
    karşılaştırmadan ÖNCE `İ`->`i`, `I`->`ı` (doğru Türkçe küçük harf
    karşılıkları) çevirip SONRA `.lower()` çağırmak bu tuzağı yapısal
    olarak imkansız kılar."""
    return metin.replace("İ", "i").replace("I", "ı").lower()


# ---------------------------------------------------------------------
# 1. Kategori sınıflandırması — anahtar kelime eşleştirmesi
# ---------------------------------------------------------------------

# Sıra ÖNEMLİDİR: bir dosya birden fazla kategoriye uyabilir, İLK eşleşen
# kategori kazanır (daha SPESİFİK kategoriler listenin BAŞINDA).
KATEGORI_ANAHTAR_KELIMELERI: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("çoklu-kart", ("coklu_kart", "sistem_orkestratoru", "arayuz_sozlesmesi", "device_tree")),
    ("HDI", ("hdi_stackup", "via_siniflandirma", "mikrovia", "blind_buried")),
    ("empedans/SI", ("empedans", "impedance", "skew", "highspeed", "high_speed", "diferansiyel",
                      "diff_pair", "openems")),
    ("termal", ("termal", "isil", "sicaklik", "junction", "soguturucu")),
    ("PDN/güç", ("pdn", "_guc_", "guc_butcesi", "power", "decoupling", "pmic", "ray_tahsisi",
                 "akim_yogunlugu")),
    ("yerleşim/placement", ("yerlesim", "placement", "kumeleri_bul")),
    ("routing", ("router", "_rota_", "routing", "izgara_a_yildiz", "carpisma_radari",
                 "topolojik_router", "kurtarma_motoru")),
    ("DFM/üretim", ("dfm", "uretim_", "gerber", "kibot", "jlcpcb", "ipc7351", "ipc6012",
                     "ipc2152", "ipc2221", "ipc_dru", "footprint", "panelizasyon")),
    ("test/doğrulama", ("dogrulama", "verify", "_kontrolu", "kontrolu_", "drc", "erc",
                         "checker", "hata_hafizasi", "karar_birimleri")),
)


def kategori_belirle(dosya_adi: str, kaynak: str) -> str:
    """Önce dosya ADINDA, bulunamazsa dosya İÇERİĞİNDE anahtar kelime arar.
    Hiçbiri eşleşmezse "diğer" döner — uydurma bir kategori ATANMAZ."""
    ad_kucuk = turkce_kucuk_harfe_cevir(dosya_adi)
    for kategori, anahtarlar in KATEGORI_ANAHTAR_KELIMELERI:
        if any(a in ad_kucuk for a in anahtarlar):
            return kategori
    kaynak_kucuk = turkce_kucuk_harfe_cevir(kaynak)
    for kategori, anahtarlar in KATEGORI_ANAHTAR_KELIMELERI:
        if any(a in kaynak_kucuk for a in anahtarlar):
            return kategori
    return "diğer"


# ---------------------------------------------------------------------
# 2. Durum tespiti — önce marker, yoksa kaba otomatik tahmin
# ---------------------------------------------------------------------

_ISKELET_MARKERLARI = ("iskelet",)
_KISMEN_MARKERLARI = ("tam uygulama değil", "tam uygulama degil", "kapsam dışı", "kapsam disi")


def durum_belirle(kaynak: str) -> Tuple[str, str]:
    """Döner: (durum, durum_kaynagi). `durum_kaynagi` "marker" ise dosyanın
    KENDİ docstring'inde açık bir işaretleyici ifade bulundu demektir
    (güvenilir); "otomatik_tahmin" ise fonksiyon sayısı/uzunluğa göre KABA
    bir kestirimdir (elle doğrulanmalı)."""
    kaynak_kucuk = turkce_kucuk_harfe_cevir(kaynak)
    if any(m in kaynak_kucuk for m in _ISKELET_MARKERLARI):
        return "İskelet", "marker"
    if any(m in kaynak_kucuk for m in _KISMEN_MARKERLARI):
        return "Kısmen var", "marker"

    ust_seviye_def_sayisi = len(re.findall(r"^(?:def|class)\s+\w+", kaynak, re.M))
    toplam_satir = len(kaynak.splitlines())

    if ust_seviye_def_sayisi <= 2 and toplam_satir < 80:
        return "İskelet", "otomatik_tahmin"
    if toplam_satir < 150:
        return "Kısmen var", "otomatik_tahmin"
    return "Var", "otomatik_tahmin"


# ---------------------------------------------------------------------
# 3. Özet çıkarımı — docstring'in ilk 1-2 cümlesi
# ---------------------------------------------------------------------


def ozet_cikar(kaynak: str) -> str:
    try:
        agac = ast.parse(kaynak)
        docstring = ast.get_docstring(agac)
    except SyntaxError:
        return "(docstring okunamadı — sözdizimi hatası)"
    if not docstring:
        return "(docstring yok)"

    # Başlık/ayraç satırlarını (===...=== gibi) ve boş satırları atla,
    # gerçek nesir metnin başladığı ilk paragrafı bul.
    paragraflar = [p.strip() for p in docstring.split("\n\n") if p.strip()]
    for p in paragraflar:
        temiz = " ".join(
            satir for satir in p.splitlines()
            if satir.strip() and not re.fullmatch(r"[=\-]{3,}", satir.strip())
        ).strip()
        if temiz:
            cumleler = re.split(r"(?<=[.!?])\s+", temiz)
            return " ".join(cumleler[:2])[:400]
    return "(docstring yalnızca başlık/ayraç içeriyor)"


# ---------------------------------------------------------------------
# 4. Gerçek/Mock şüpheli desen taraması
# ---------------------------------------------------------------------

_SUPHELI_ANAHTAR_KELIMELER = ("mock", "sahte", "taklit", "hardcode", "placeholder")
_SABIT_DONEN_IF_DESENI = re.compile(r"if\s+[^\n:]+==[^\n:]+:\s*\n\s*return\s+[\d.\"']")


def gercek_mi_mock_mu_tara(kaynak: str) -> Tuple[str, List[str]]:
    """`mock`/`sahte`/`taklit` gibi anahtar kelimeler VEYA tekrarlanan
    sabit-değer-dönen if/elif zincirleri bulursa "KONTROL EDİLMELİ" +
    bulguları döner. Bu bir KESİN karar DEĞİLDİR — ör. "taklit_pcbnew"
    test yardımcı sınıfı adı test dosyalarında MEŞRUDUR (bu fonksiyon
    zaten test_*.py dosyalarına ÇAĞRILMAZ, bkz. `repo_taramasi_yap()`)."""
    bulgular: List[str] = []
    kaynak_kucuk = kaynak.lower()
    for kw in _SUPHELI_ANAHTAR_KELIMELER:
        if kw in kaynak_kucuk:
            bulgular.append(f"'{kw}' kelimesi geçiyor")

    sabit_donusler = _SABIT_DONEN_IF_DESENI.findall(kaynak)
    if len(sabit_donusler) >= 2:
        bulgular.append(f"{len(sabit_donusler)} adet sabit-değer-dönen if/elif zinciri")

    if bulgular:
        return "KONTROL EDİLMELİ", bulgular
    return "işaret yok (otomatik tarama)", []


# ---------------------------------------------------------------------
# 5. Birleşik Araç Zinciri Uyumu (çiz -> simüle/hesapla -> JSON geri bildir)
# ---------------------------------------------------------------------

_JSON_DESENI = re.compile(r"\bimport json\b|\bjson\.load\b|\bjson\.dump\b")
_DONGU_HEDEF_DESENI = re.compile(
    r"\bwhile\b[^\n]*\n(?:[^\n]*\n){0,15}?[^\n]*(?:hedef|target|yakinsa|converge|max_deneme|maks_iterasyon)",
    re.I,
)
_HESAPLA_CIZ_DESENI = re.compile(
    r"\bpcbnew\b|\bhesapla\b|\bcoz\b|\brout[ei]\b|VECTOR2I|empedans",
    re.I,
)


def arac_zinciri_uyumu_tahmin_et(kaynak: str) -> str:
    json_var = bool(_JSON_DESENI.search(kaynak))
    dongu_var = bool(_DONGU_HEDEF_DESENI.search(kaynak))
    hesapla_ciz_var = bool(_HESAPLA_CIZ_DESENI.search(kaynak))

    if json_var and dongu_var and hesapla_ciz_var:
        return "Tam uyumlu"
    if hesapla_ciz_var or json_var:
        return "Kısmen uyumlu"
    return "İlgisiz"


# ---------------------------------------------------------------------
# 6. Son değişiklik bilgisi (git log)
# ---------------------------------------------------------------------


def son_degisiklik_bilgisi(dosya: Path, repo_kok: Path) -> Tuple[str, str]:
    """`git log -1` ile (tarih, mesaj) döner. Git geçmişi yoksa/başarısız
    olursa ("bilinmiyor", "") döner — uydurma bir tarih ÜRETİLMEZ."""
    try:
        sonuc = subprocess.run(
            ["git", "log", "-1", "--format=%ad|%s", "--date=short", "--", dosya.name],
            cwd=repo_kok, capture_output=True, text=True, timeout=15, check=False,
        )
    except (OSError, subprocess.TimeoutExpired):
        return "bilinmiyor", ""
    if sonuc.returncode != 0 or not sonuc.stdout.strip():
        return "bilinmiyor (commit edilmemiş olabilir)", ""
    tarih, _, mesaj = sonuc.stdout.strip().partition("|")
    return tarih, mesaj


# ---------------------------------------------------------------------
# 7. Test durumu — GERÇEK pytest çalıştırması (junit-xml üzerinden)
# ---------------------------------------------------------------------


def tum_testleri_calistir_ve_sonuclari_topla(
    repo_kok: Path, zaman_asimi_s: int = 900,
) -> Dict[str, Tuple[int, int]]:
    """Tüm test suite'ini BİR KEZ `pytest --junit-xml` ile çalıştırır, her
    test dosyası (classname) için (geçen, toplam) test sayısını döner.

    GERÇEK bir çalıştırmadır — sabit/varsayılan bir "hepsi geçti" hiçbir
    zaman UYDURULMAZ. `pytest`/xml üretilemezse (ör. ortam sorunu) boş
    sözlük döner, çağıran taraf bunu "ölçülemedi" olarak yorumlamalı.
    """
    with tempfile.TemporaryDirectory() as tmp:
        xml_yolu = Path(tmp) / "sonuclar.xml"
        try:
            subprocess.run(
                [sys.executable, "-m", "pytest", "-q", "--tb=no", f"--junit-xml={xml_yolu}"],
                cwd=repo_kok, timeout=zaman_asimi_s, capture_output=True, check=False,
            )
        except subprocess.TimeoutExpired:
            return {}
        if not xml_yolu.exists():
            return {}

        try:
            agac = ET.parse(xml_yolu)
        except ET.ParseError:
            return {}

        sonuclar: Dict[str, Tuple[int, int]] = {}
        for testcase in agac.getroot().iter("testcase"):
            classname = testcase.get("classname", "") or ""
            dosya_adi = classname.split(".")[-1] if classname else ""
            if not dosya_adi:
                continue
            gecen, toplam = sonuclar.get(dosya_adi, (0, 0))
            toplam += 1
            basarisiz = (
                testcase.find("failure") is not None
                or testcase.find("error") is not None
            )
            if not basarisiz:
                gecen += 1
            sonuclar[dosya_adi] = (gecen, toplam)
        return sonuclar


# ---------------------------------------------------------------------
# 8. Tek modül kaydı + toplu tarama
# ---------------------------------------------------------------------


@dataclass
class ModulKaydi:
    modul_adi: str
    kategori: str
    durum: str
    durum_kaynagi: str
    ozet: str
    test_dosyasi_var_mi: bool
    test_fonksiyon_sayisi: int
    test_durumu: str  # "gecti" | "basarisiz" | "test_yok" | "olculemedi"
    gercek_mi_mock_mu: str
    mock_bulgulari: List[str] = field(default_factory=list)
    son_degisiklik_tarihi: str = "bilinmiyor"
    son_degisiklik_mesaji: str = ""
    arac_zinciri_uyumu: str = "İlgisiz"


def modul_kaydi_olustur(
    dosya: Path, repo_kok: Path, test_sonuclari: Optional[Dict[str, Tuple[int, int]]],
) -> ModulKaydi:
    kaynak = dosya.read_text(encoding="utf-8")

    kategori = kategori_belirle(dosya.name, kaynak)
    durum, durum_kaynagi = durum_belirle(kaynak)
    ozet = ozet_cikar(kaynak)
    gercek_mock, mock_bulgulari = gercek_mi_mock_mu_tara(kaynak)
    arac_uyumu = arac_zinciri_uyumu_tahmin_et(kaynak)
    tarih, mesaj = son_degisiklik_bilgisi(dosya, repo_kok)

    test_dosyasi = repo_kok / f"test_{dosya.stem}.py"
    test_var = test_dosyasi.exists()
    test_fonksiyon_sayisi = 0
    if test_var:
        test_kaynak = test_dosyasi.read_text(encoding="utf-8")
        test_fonksiyon_sayisi = len(re.findall(r"^\s*def test_\w+", test_kaynak, re.M))

    if not test_var:
        test_durumu = "test_yok"
    elif test_sonuclari is None:
        test_durumu = "olculemedi"
    else:
        anahtar = f"test_{dosya.stem}"
        if anahtar not in test_sonuclari:
            test_durumu = "olculemedi"
        else:
            gecen, toplam = test_sonuclari[anahtar]
            test_durumu = "gecti" if (toplam > 0 and gecen == toplam) else "basarisiz"

    return ModulKaydi(
        modul_adi=dosya.name,
        kategori=kategori,
        durum=durum,
        durum_kaynagi=durum_kaynagi,
        ozet=ozet,
        test_dosyasi_var_mi=test_var,
        test_fonksiyon_sayisi=test_fonksiyon_sayisi,
        test_durumu=test_durumu,
        gercek_mi_mock_mu=gercek_mock,
        mock_bulgulari=mock_bulgulari,
        son_degisiklik_tarihi=tarih,
        son_degisiklik_mesaji=mesaj,
        arac_zinciri_uyumu=arac_uyumu,
    )


def modul_dosyalarini_bul(repo_kok: Path) -> List[Path]:
    """Repo kökündeki TÜM `.py` dosyalarını (test_*.py HARİÇ) döner —
    alt dizinlere İNMEZ (proje düz/flat modül yapısı kullanır)."""
    return sorted(
        p for p in repo_kok.glob("*.py")
        if not p.name.startswith("test_")
    )


def repo_taramasi_yap(
    repo_kok: Path, test_sonuclari: Optional[Dict[str, Tuple[int, int]]] = None,
) -> List[ModulKaydi]:
    dosyalar = modul_dosyalarini_bul(repo_kok)
    return [modul_kaydi_olustur(d, repo_kok, test_sonuclari) for d in dosyalar]


# ---------------------------------------------------------------------
# 9. Excel çıktısı
# ---------------------------------------------------------------------

_DURUM_RENKLERI = {
    "Var": "C6EFCE",
    "Kısmen var": "FFEB9C",
    "İskelet": "D9D9D9",
    "KONTROL EDİLMELİ": "FFC7CE",
}

_DETAY_BASLIKLARI = (
    "Modül", "Kategori", "Durum", "Durum Kaynağı", "Özet",
    "Test Var mı", "Test Fonksiyon Sayısı", "Test Durumu",
    "Gerçek/Mock", "Mock Bulguları", "Son Değişiklik Tarihi",
    "Son Değişiklik Mesajı", "Araç Zinciri Uyumu",
)


def excel_uret(kayitlar: List[ModulKaydi], cikti_yolu: Path) -> None:
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()

    ozet_ws = wb.active
    ozet_ws.title = "Özet"
    ozet_ws.append([f"YETENEK ENVANTERİ — son güncelleme: {datetime.now():%Y-%m-%d %H:%M}"])
    ozet_ws.append([])
    ozet_ws.append([
        "UYARI: 'Durum Kaynağı' sütunu 'otomatik_tahmin' olan satırlar KABA "
        "bir kestirimdir (fonksiyon sayısı/uzunluk), elle doğrulanmamıştır. "
        "Kritik bir karardan önce ilgili dosyayı ELLE kontrol edin."
    ])
    ozet_ws.append([])
    ozet_ws.append(["Kategori", "Var", "Kısmen var", "İskelet", "Toplam"])

    kategori_sayimlari: Dict[str, Dict[str, int]] = {}
    for k in kayitlar:
        d = kategori_sayimlari.setdefault(k.kategori, {"Var": 0, "Kısmen var": 0, "İskelet": 0})
        d[k.durum] = d.get(k.durum, 0) + 1

    for kategori in sorted(kategori_sayimlari):
        d = kategori_sayimlari[kategori]
        toplam = d["Var"] + d["Kısmen var"] + d["İskelet"]
        ozet_ws.append([kategori, d["Var"], d["Kısmen var"], d["İskelet"], toplam])

    ozet_ws.append([])
    toplam_modul = len(kayitlar)
    kontrol_edilmesi_gereken = sum(1 for k in kayitlar if k.gercek_mi_mock_mu == "KONTROL EDİLMELİ")
    testsiz = sum(1 for k in kayitlar if k.test_durumu == "test_yok")
    ozet_ws.append([f"Toplam modül: {toplam_modul}"])
    ozet_ws.append([f"'KONTROL EDİLMELİ' işaretli modül: {kontrol_edilmesi_gereken}"])
    ozet_ws.append([f"Test dosyası OLMAYAN modül: {testsiz}"])

    detay_ws = wb.create_sheet("Detay")
    detay_ws.append(list(_DETAY_BASLIKLARI))
    for k in kayitlar:
        detay_ws.append([
            k.modul_adi, k.kategori, k.durum, k.durum_kaynagi, k.ozet,
            "Evet" if k.test_dosyasi_var_mi else "Hayır", k.test_fonksiyon_sayisi,
            k.test_durumu, k.gercek_mi_mock_mu, "; ".join(k.mock_bulgulari),
            k.son_degisiklik_tarihi, k.son_degisiklik_mesaji, k.arac_zinciri_uyumu,
        ])

    durum_sutunu = _DETAY_BASLIKLARI.index("Durum") + 1
    mock_sutunu = _DETAY_BASLIKLARI.index("Gerçek/Mock") + 1
    for satir_no in range(2, len(kayitlar) + 2):
        durum_hucre = detay_ws.cell(row=satir_no, column=durum_sutunu)
        renk = _DURUM_RENKLERI.get(durum_hucre.value)
        if renk:
            durum_hucre.fill = PatternFill(start_color=renk, end_color=renk, fill_type="solid")
        mock_hucre = detay_ws.cell(row=satir_no, column=mock_sutunu)
        renk = _DURUM_RENKLERI.get(mock_hucre.value)
        if renk:
            mock_hucre.fill = PatternFill(start_color=renk, end_color=renk, fill_type="solid")

    detay_ws.auto_filter.ref = detay_ws.dimensions

    cikti_yolu.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(cikti_yolu))


# ---------------------------------------------------------------------
# 10. CLI
# ---------------------------------------------------------------------


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--repo-kok", dest="repo_kok", default=".", help="repo kök dizini (varsayılan: .)")
    ap.add_argument("--cikti", default="YETENEK_ENVANTERI.xlsx", help="çıktı .xlsx dosya adı/yolu")
    ap.add_argument(
        "--testleri-atla", dest="testleri_atla", action="store_true",
        help="tam suite'i çalıştırmayı atla (hızlı ama Test Durumu='olculemedi' kalır)",
    )
    a = ap.parse_args(argv)

    repo_kok = Path(a.repo_kok).resolve()
    test_sonuclari = None if a.testleri_atla else tum_testleri_calistir_ve_sonuclari_topla(repo_kok)
    kayitlar = repo_taramasi_yap(repo_kok, test_sonuclari)

    cikti_yolu = Path(a.cikti)
    if not cikti_yolu.is_absolute():
        cikti_yolu = repo_kok / cikti_yolu
    excel_uret(kayitlar, cikti_yolu)

    print(f"{len(kayitlar)} modül tarandı, envanter yazıldı: {cikti_yolu}")
    kontrol_edilmesi_gereken = [k.modul_adi for k in kayitlar if k.gercek_mi_mock_mu == "KONTROL EDİLMELİ"]
    if kontrol_edilmesi_gereken:
        print(f"  KONTROL EDİLMELİ ({len(kontrol_edilmesi_gereken)}): {', '.join(kontrol_edilmesi_gereken)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

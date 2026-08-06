"""yetenek_envanteri_uret.py için test suite.
Çalıştırmak için: pytest -v test_yetenek_envanteri_uret.py
"""

from __future__ import annotations

import openpyxl
import pytest

from yetenek_envanteri_uret import (
    ModulKaydi,
    arac_zinciri_uyumu_tahmin_et,
    durum_belirle,
    excel_uret,
    gercek_mi_mock_mu_tara,
    kategori_belirle,
    modul_dosyalarini_bul,
    modul_kaydi_olustur,
    ozet_cikar,
    repo_taramasi_yap,
    tum_testleri_calistir_ve_sonuclari_topla,
    turkce_kucuk_harfe_cevir,
)


# ------------------------------------------------------------------
# 0. turkce_kucuk_harfe_cevir — FAULT-INJECTION: bu regresyon `str.lower()`
# ile İLK yazımda gerçekten bulundu ("DEĞİL".lower() "değil" ile eşleşmiyordu,
# Windows konsolunda UnicodeEncodeError bile fırlatıyordu).
# ------------------------------------------------------------------

class TestTurkceKucukHarfeCevir:
    def test_noktali_buyuk_i_duz_i_ye_doner(self):
        assert turkce_kucuk_harfe_cevir("İSKELET") == "iskelet"

    def test_stdlib_lower_ile_esitsizligi_kanitlar(self):
        # str.lower() 'İ'yi 'i' + BİRLEŞTİRİCİ NOKTA (U+0307)'a çevirir —
        # bu satır, düzeltmenin GERÇEKTEN gerekli olduğunun kanıtıdır.
        assert "DEĞİL".lower() != "değil"
        assert turkce_kucuk_harfe_cevir("DEĞİL") == "değil"

    def test_noktasiz_buyuk_i_de_duz_i_ye_doner(self):
        assert turkce_kucuk_harfe_cevir("KAPSAM DIŞI") == "kapsam dışı"


# ------------------------------------------------------------------
# 1. kategori_belirle
# ------------------------------------------------------------------

class TestKategoriBelirle:
    def test_dosya_adindan_eslesir(self):
        assert kategori_belirle("coklu_kart_sozlesme_kontrolu.py", "") == "çoklu-kart"
        assert kategori_belirle("pcb_highspeed_escape.py", "") == "empedans/SI"

    def test_icerikten_eslesir_ad_eslesmezse(self):
        assert kategori_belirle("genel_yardimci.py", "bu dosya empedans hesaplar") == "empedans/SI"

    def test_hicbiri_eslesmezse_diger(self):
        assert kategori_belirle("tamamen_alakasiz_isim.py", "hiçbir anahtar kelime yok burada") == "diğer"


# ------------------------------------------------------------------
# 2. durum_belirle
# ------------------------------------------------------------------

class TestDurumBelirle:
    def test_iskelet_markeri_yakalanir(self):
        kaynak = '"""Bu dosya sadece bir İSKELET içerir, gerçek mantık yok."""\n'
        assert durum_belirle(kaynak) == ("İskelet", "marker")

    def test_kismen_var_markeri_yakalanir(self):
        kaynak = '"""Bu fonksiyon TAM UYGULAMA DEĞİL, sadece analitik yaklaşım."""\n'
        assert durum_belirle(kaynak) == ("Kısmen var", "marker")

    def test_kapsam_disi_markeri_kismen_var_sayilir(self):
        kaynak = '"""Bu özellik kapsam dışı bırakıldı."""\n'
        assert durum_belirle(kaynak) == ("Kısmen var", "marker")

    def test_marker_yoksa_kisa_dosya_iskelet_tahmini(self):
        kaynak = '"""Kısa bir dosya."""\ndef tek_fonksiyon():\n    pass\n'
        durum, kaynak_tipi = durum_belirle(kaynak)
        assert durum == "İskelet"
        assert kaynak_tipi == "otomatik_tahmin"

    def test_marker_yoksa_uzun_dosya_var_tahmini(self):
        gövde = "\n".join(f"def fonksiyon_{i}():\n    return {i}\n" for i in range(60))
        kaynak = '"""Uzun bir dosya."""\n' + gövde
        durum, kaynak_tipi = durum_belirle(kaynak)
        assert durum == "Var"
        assert kaynak_tipi == "otomatik_tahmin"


# ------------------------------------------------------------------
# 3. ozet_cikar
# ------------------------------------------------------------------

class TestOzetCikar:
    def test_ilk_iki_cumleyi_alir(self):
        kaynak = '"""Bu ilk cümle. Bu ikinci cümle. Bu üçüncü cümle asla dönmemeli."""\n'
        ozet = ozet_cikar(kaynak)
        assert "Bu ilk cümle." in ozet
        assert "Bu ikinci cümle." in ozet
        assert "üçüncü cümle" not in ozet

    def test_ayrac_satirlarini_atlar(self):
        kaynak = '"""\nbaslik.py\n===========\nGerçek açıklama burada başlıyor."""\n'
        ozet = ozet_cikar(kaynak)
        assert "===" not in ozet
        assert "Gerçek açıklama" in ozet

    def test_docstring_yoksa_belirtir(self):
        assert ozet_cikar("x = 1\n") == "(docstring yok)"

    def test_sozdizimi_hatasinda_cokmez(self):
        assert "okunamadı" in ozet_cikar("def bozuk(:\n")


# ------------------------------------------------------------------
# 4. gercek_mi_mock_mu_tara
# ------------------------------------------------------------------

class TestGercekMiMockMuTara:
    def test_supheli_kelime_yakalanir(self):
        durum, bulgular = gercek_mi_mock_mu_tara("def f():\n    return MOCK_DEGER  # mock veri\n")
        assert durum == "KONTROL EDİLMELİ"
        assert any("mock" in b for b in bulgular)

    def test_sabit_donen_zincir_yakalanir(self):
        kaynak = (
            "def f(x):\n"
            "    if x == 1:\n        return 100\n"
            "    if x == 2:\n        return 200\n"
        )
        durum, bulgular = gercek_mi_mock_mu_tara(kaynak)
        assert durum == "KONTROL EDİLMELİ"
        assert any("if/elif" in b for b in bulgular)

    def test_temiz_dosya_isaret_yok(self):
        kaynak = "def gercek_hesap(genislik, yukseklik):\n    return genislik * yukseklik\n"
        durum, bulgular = gercek_mi_mock_mu_tara(kaynak)
        assert durum == "işaret yok (otomatik tarama)"
        assert bulgular == []


# ------------------------------------------------------------------
# 5. arac_zinciri_uyumu_tahmin_et
# ------------------------------------------------------------------

class TestAracZinciriUyumu:
    def test_tam_uyumlu_json_dongu_ve_hesap_birlikte(self):
        kaynak = (
            "import json\n"
            "def uygula(hedef):\n"
            "    while abs(mevcut - hedef) > tolerans:\n"
            "        mevcut = empedans_hesapla(genislik)\n"
            "        json.dump({'target': hedef}, f)\n"
        )
        assert arac_zinciri_uyumu_tahmin_et(kaynak) == "Tam uyumlu"

    def test_kismen_uyumlu_sadece_hesap(self):
        kaynak = "def empedans_hesapla(w, h):\n    return w / h\n"
        assert arac_zinciri_uyumu_tahmin_et(kaynak) == "Kısmen uyumlu"

    def test_ilgisiz_ne_hesap_ne_json(self):
        kaynak = "def dosya_kopyala(a, b):\n    import shutil\n    shutil.copy(a, b)\n"
        assert arac_zinciri_uyumu_tahmin_et(kaynak) == "İlgisiz"


# ------------------------------------------------------------------
# 6. modul_dosyalarini_bul / repo_taramasi_yap / modul_kaydi_olustur
# ------------------------------------------------------------------

class TestRepoTaramasi:
    def _sahte_repo_kur(self, tmp_path):
        (tmp_path / "gercek_modul.py").write_text(
            '"""gercek_modul.py\n=====\nGerçekten çalışan bir hesaplayıcı."""\n'
            + "\n".join(f"def f{i}():\n    return {i}\n" for i in range(60)),
            encoding="utf-8",
        )
        (tmp_path / "test_gercek_modul.py").write_text(
            "def test_a():\n    assert True\n\ndef test_b():\n    assert True\n",
            encoding="utf-8",
        )
        return tmp_path

    def test_test_dosyalari_haric_tutulur(self, tmp_path):
        self._sahte_repo_kur(tmp_path)
        dosyalar = modul_dosyalarini_bul(tmp_path)
        assert [d.name for d in dosyalar] == ["gercek_modul.py"]

    def test_modul_kaydi_test_yok_durumu(self, tmp_path):
        (tmp_path / "testsiz_modul.py").write_text('"""Testsiz modül."""\ndef f():\n    pass\n', encoding="utf-8")
        kayit = modul_kaydi_olustur(tmp_path / "testsiz_modul.py", tmp_path, test_sonuclari=None)
        assert kayit.test_dosyasi_var_mi is False
        assert kayit.test_durumu == "test_yok"

    def test_modul_kaydi_test_sonuclari_gecti(self, tmp_path):
        self._sahte_repo_kur(tmp_path)
        kayit = modul_kaydi_olustur(
            tmp_path / "gercek_modul.py", tmp_path,
            test_sonuclari={"test_gercek_modul": (2, 2)},
        )
        assert kayit.test_dosyasi_var_mi is True
        assert kayit.test_fonksiyon_sayisi == 2
        assert kayit.test_durumu == "gecti"

    def test_modul_kaydi_test_sonuclari_basarisiz(self, tmp_path):
        self._sahte_repo_kur(tmp_path)
        kayit = modul_kaydi_olustur(
            tmp_path / "gercek_modul.py", tmp_path,
            test_sonuclari={"test_gercek_modul": (1, 2)},
        )
        assert kayit.test_durumu == "basarisiz"

    def test_modul_kaydi_test_sonuclari_verilmezse_olculemedi(self, tmp_path):
        self._sahte_repo_kur(tmp_path)
        kayit = modul_kaydi_olustur(tmp_path / "gercek_modul.py", tmp_path, test_sonuclari=None)
        assert kayit.test_durumu == "olculemedi"

    def test_repo_taramasi_tum_dosyalari_isler(self, tmp_path):
        self._sahte_repo_kur(tmp_path)
        kayitlar = repo_taramasi_yap(tmp_path, test_sonuclari={"test_gercek_modul": (2, 2)})
        assert len(kayitlar) == 1
        assert kayitlar[0].modul_adi == "gercek_modul.py"
        assert kayitlar[0].durum == "Var"


# ------------------------------------------------------------------
# 7. tum_testleri_calistir_ve_sonuclari_topla — GERÇEK pytest çalıştırması
# ------------------------------------------------------------------

class TestGercekPytestCalistirma:
    def test_gercek_pass_ve_fail_dogru_ayristirilir(self, tmp_path):
        """Sahte, minik bir 'proje' üzerinde GERÇEK bir pytest alt-süreci
        çalıştırıp junit-xml ayrıştırmasının doğru olduğunu kanıtlar —
        mock DEĞİL, gerçek bir pytest çıktısı okunur."""
        (tmp_path / "test_karisik.py").write_text(
            "def test_gecen():\n    assert True\n\n"
            "def test_basarisiz():\n    assert False\n",
            encoding="utf-8",
        )
        sonuclar = tum_testleri_calistir_ve_sonuclari_topla(tmp_path, zaman_asimi_s=60)
        assert sonuclar["test_karisik"] == (1, 2)


# ------------------------------------------------------------------
# 8. excel_uret — üretilen .xlsx gerçekten açılabilir mi
# ------------------------------------------------------------------

class TestExcelUret:
    def test_uretilen_xlsx_beklenen_sayfa_ve_sutun_yapisinda(self, tmp_path):
        kayitlar = [
            ModulKaydi(
                modul_adi="a.py", kategori="routing", durum="Var", durum_kaynagi="marker",
                ozet="Bir özet.", test_dosyasi_var_mi=True, test_fonksiyon_sayisi=3,
                test_durumu="gecti", gercek_mi_mock_mu="işaret yok (otomatik tarama)",
                mock_bulgulari=[], son_degisiklik_tarihi="2026-08-06",
                son_degisiklik_mesaji="ilk commit", arac_zinciri_uyumu="Kısmen uyumlu",
            ),
            ModulKaydi(
                modul_adi="b.py", kategori="diğer", durum="İskelet", durum_kaynagi="otomatik_tahmin",
                ozet="Başka özet.", test_dosyasi_var_mi=False, test_fonksiyon_sayisi=0,
                test_durumu="test_yok", gercek_mi_mock_mu="KONTROL EDİLMELİ",
                mock_bulgulari=["'mock' kelimesi geçiyor"], son_degisiklik_tarihi="bilinmiyor",
                son_degisiklik_mesaji="", arac_zinciri_uyumu="İlgisiz",
            ),
        ]
        cikti = tmp_path / "YETENEK_ENVANTERI.xlsx"

        excel_uret(kayitlar, cikti)

        assert cikti.exists()
        wb = openpyxl.load_workbook(cikti)
        assert "Özet" in wb.sheetnames
        assert "Detay" in wb.sheetnames

        detay = wb["Detay"]
        basliklar = [c.value for c in detay[1]]
        assert basliklar[:3] == ["Modül", "Kategori", "Durum"]
        assert detay.max_row == 3  # başlık + 2 kayıt
        assert detay.cell(row=2, column=1).value == "a.py"
        assert detay.cell(row=3, column=1).value == "b.py"

        # Renk kodlama uygulanmış mı (dolgu rengi ayarlanmış)
        durum_hucresi_iskelet = detay.cell(row=3, column=3)
        assert durum_hucresi_iskelet.fill.start_color.rgb not in (None, "00000000")

        ozet = wb["Özet"]
        ozet_metni = "\n".join(str(c[0].value) for c in ozet.iter_rows() if c[0].value)
        assert "Toplam modül: 2" in ozet_metni
